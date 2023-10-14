import datetime

import openpyxl
from openpyxl.utils import get_column_letter

def update_data_to_excel(excel_file, Data):
    
    # Load the excel file
    workbook = openpyxl.load_workbook(excel_file)

    # Select the Worksheet
    worksheet = workbook['Sheet1']

    todayDate = datetime.date.today().strftime('%d-%m-%Y')

    # Add a new column with today's date
    new_column_letter = get_column_letter(worksheet.max_column + 1)
    new_column_name = todayDate
    worksheet[new_column_letter + '1'] = new_column_name

    # Add attendance data to the created column of each student
    for row in range(2, worksheet.max_row + 1):
        student_USN = worksheet.cell(row=row, column=1).value
        if student_USN in Data:
            attendance_status = Data[student_USN]
            worksheet[new_column_letter + str(row)] = attendance_status

    workbook.save(excel_file)
